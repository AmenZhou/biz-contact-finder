import openpyxl

wb = openpyxl.load_workbook('Ads team Xiulian_Tracker_2025-Ads team-1.xlsx')
print(f"Sheet names: {wb.sheetnames}")
print()

for sheet_name in wb.sheetnames:
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet_name}")
    print('='*60)
    ws = wb[sheet_name]

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(row)
        if i > 100:  # Limit output
            print(f"... (showing first 100 rows)")
            break
