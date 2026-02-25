#!/usr/bin/env python3
"""
Generate KMZ for EDDM carrier routes listed in DTD-2026-2.xlsx.

Reads the xlsx file, fetches route geometry from the USPS EDDM ArcGIS API,
and outputs a KMZ with color-coded routes and demographic popups.

Usage:
    python scripts/dtd/01_export_to_kmz.py
"""

import time
import zipfile
from collections import defaultdict
from pathlib import Path

import openpyxl
import requests
from simplekml import Kml

PROJECT_ROOT = Path(__file__).parent.parent.parent
XLSX_PATH = PROJECT_ROOT / "DTD-2026-2.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "data" / "dtd"
OUTPUT_NAME = "dtd_2026_2_routes.kmz"

EDDM_API_URL = "https://gis.usps.com/arcgis/rest/services/EDDM/selectZIP/GPServer/routes/execute"

ROUTE_COLORS = [
    "ff0000ff",  # Red
    "ffff0000",  # Blue
    "ff00cc00",  # Green
    "ff00a5ff",  # Orange
    "ffff00ff",  # Magenta
    "ff00ffff",  # Yellow
    "ffaa00ff",  # Pink
    "ff00ffaa",  # Lime
    "ffccaa00",  # Teal
    "ff5050ff",  # Salmon
    "ffff5050",  # Light Blue
    "ff50ff50",  # Light Green
]


def load_xlsx(path):
    """Load ZIP/CRID pairs and demographic data from the xlsx file.

    Returns:
        dict: {(zip, crid): {demographic fields...}}
        grouped: {zip: set of crids}
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    routes = {}
    grouped = defaultdict(set)

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        zipcode = str(data.get("ZIP", "")).strip()
        crid = str(data.get("CRID", "")).strip()
        if not zipcode or not crid:
            continue

        grouped[zipcode].add(crid)
        routes[(zipcode, crid)] = {
            "city": str(data.get("City", "")).strip(),
            "state": str(data.get("State", "")).strip(),
            "segment": str(data.get("Segment", "")).strip(),
            "sfdu": data.get("SFDU", 0) or 0,
            "mfdu": data.get("MFDU", 0) or 0,
            "trailers": data.get("Trailers", 0) or 0,
            "sub_total": data.get("Sub Total", 0) or 0,
            "bus": data.get("Bus", 0) or 0,
            "total": data.get("Total", 0) or 0,
            "names": data.get("Names", 0) or 0,
            "median_income": data.get("Median Income"),
            "median_home_value": data.get("Median Home Value"),
            "median_age": data.get("Median Age"),
            "phwc": str(data.get("PHWC", "")).strip(),
            "sat": str(data.get("Sat", "")).strip(),
            "dfo": str(data.get("DFO", "")).strip(),
            "pct_black": str(data.get("PCT Black", "")).strip(),
            "pct_asian": str(data.get("PCT Asian", "")).strip(),
            "pct_hispanic": str(data.get("PCT Hispanic", "")).strip(),
        }

    wb.close()
    return routes, grouped


def fetch_routes(zipcode):
    """Fetch carrier route data from the USPS EDDM API."""
    params = {
        "f": "json",
        "env:outSR": "4326",
        "ZIP": zipcode,
        "Rte_Box": "R",
        "UserName": "EDDM",
    }
    response = requests.get(EDDM_API_URL, params=params, timeout=30)
    response.raise_for_status()
    data = response.json()

    if "error" in data:
        raise RuntimeError(f"USPS API error for {zipcode}: {data['error'].get('message', data['error'])}")

    results = data.get("results", [])
    if not results:
        return []
    return results[0].get("value", {}).get("features", [])


def build_description(xlsx_data, api_attrs):
    """Build an HTML popup from xlsx demographic data + API attributes."""
    d = xlsx_data
    zipcode = api_attrs.get("ZIP_CODE", "N/A")
    facility = api_attrs.get("FACILITY_NAME", "N/A")

    income_str = f"${d['median_income']:,.0f}" if d.get("median_income") else "N/A"
    home_str = f"${d['median_home_value']:,.0f}" if d.get("median_home_value") else "N/A"
    age_str = str(d["median_age"]) if d.get("median_age") else "N/A"

    return f"""<b>Route:</b> {zipcode}-{api_attrs.get('CRID_ID', 'N/A')}<br/>
<b>City:</b> {d['city']}, {d['state']}<br/>
<b>Post Office:</b> {facility}<br/>
<hr/>
<b>SFDU:</b> {d['sfdu']:,} | <b>MFDU:</b> {d['mfdu']:,} | <b>Trailers:</b> {d['trailers']:,}<br/>
<b>Residential Sub Total:</b> {d['sub_total']:,}<br/>
<b>Business:</b> {d['bus']:,}<br/>
<b>Total:</b> {d['total']:,} | <b>Names:</b> {d['names']:,}<br/>
<hr/>
<b>Median Income:</b> {income_str}<br/>
<b>Median Home Value:</b> {home_str}<br/>
<b>Median Age:</b> {age_str}<br/>
<b>PHWC:</b> {d['phwc']}<br/>
<b>Saturday Delivery:</b> {d['sat']}<br/>
<b>DFO:</b> {d['dfo']}<br/>
<hr/>
<b>PCT Black:</b> {d['pct_black']} | <b>PCT Asian:</b> {d['pct_asian']} | <b>PCT Hispanic:</b> {d['pct_hispanic']}<br/>
<hr/>
<i>Source: USPS EDDM / DTD-2026-2.xlsx</i>"""


def main():
    if not XLSX_PATH.exists():
        print(f"Error: xlsx file not found: {XLSX_PATH}")
        return

    routes, grouped = load_xlsx(XLSX_PATH)
    total_routes = len(routes)

    print("=" * 60)
    print("DTD-2026-2 EDDM CARRIER ROUTES -> KMZ")
    print("=" * 60)
    print(f"Source: {XLSX_PATH.name}")
    print(f"Unique ZIPs: {len(grouped)}")
    print(f"Total routes: {total_routes}")
    print()

    kml = Kml()
    kml.document.name = "DTD-2026-2 Carrier Routes"
    kml.document.description = (
        f"EDDM carrier routes from DTD-2026-2 ({total_routes} routes across {len(grouped)} ZIPs)"
    )

    matched = 0
    missing = []
    global_color_idx = 0

    for zipcode in sorted(grouped.keys()):
        wanted_crids = grouped[zipcode]
        print(f"  ZIP {zipcode}: fetching ({len(wanted_crids)} routes)...", end=" ", flush=True)

        try:
            features = fetch_routes(zipcode)
        except Exception as e:
            print(f"ERROR: {e}")
            for crid in wanted_crids:
                missing.append((zipcode, crid))
            continue

        by_crid = {}
        for f in features:
            crid_id = f.get("attributes", {}).get("CRID_ID", "")
            by_crid[crid_id] = f

        zip_found = 0

        for crid in sorted(wanted_crids):
            feature = by_crid.get(crid)
            if not feature:
                missing.append((zipcode, crid))
                continue

            attrs = feature.get("attributes", {})
            geometry = feature.get("geometry", {})
            paths = geometry.get("paths", [])
            if not paths:
                missing.append((zipcode, crid))
                continue

            xlsx_data = routes[(zipcode, crid)]
            color = ROUTE_COLORS[global_color_idx % len(ROUTE_COLORS)]
            global_color_idx += 1

            label = f"{zipcode}-{crid} ({xlsx_data['total']:,} total)"
            mg = kml.newmultigeometry(name=label)
            mg.description = build_description(xlsx_data, attrs)
            mg.style.linestyle.color = color
            mg.style.linestyle.width = 4
            for path in paths:
                coords = [(pt[0], pt[1]) for pt in path]
                mg.newlinestring().coords = coords

            matched += 1
            zip_found += 1

        print(f"found {zip_found}/{len(wanted_crids)}")
        time.sleep(0.3)

    # Save KMZ
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    kml_path = OUTPUT_DIR / OUTPUT_NAME.replace(".kmz", ".kml")
    kmz_path = OUTPUT_DIR / OUTPUT_NAME

    kml.save(str(kml_path))
    with zipfile.ZipFile(kmz_path, "w", zipfile.ZIP_DEFLATED) as kmz:
        kmz.write(kml_path, arcname="doc.kml")
    kml_path.unlink()

    kmz_size = kmz_path.stat().st_size / 1024
    print()
    print("=" * 60)
    print("EXPORT COMPLETE")
    print("=" * 60)
    print(f"Output: {kmz_path} ({kmz_size:.1f} KB)")
    print(f"Routes matched: {matched}/{total_routes}")

    if missing:
        print(f"\nMissing routes ({len(missing)}):")
        for z, c in missing:
            print(f"   {z} {c}")

    print()
    print("Import into Google My Maps:")
    print("   1. Go to https://mymaps.google.com")
    print("   2. Create a New Map -> Import -> Upload the .kmz file")
    print("=" * 60)


if __name__ == "__main__":
    main()
